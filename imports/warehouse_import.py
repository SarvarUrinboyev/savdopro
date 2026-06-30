"""
SavdoPRO — safe Warehouse Excel importer (upsert by code, never deletes).

SAFETY
------
- Refuses to run unless  ALLOW_WAREHOUSE_IMPORT=true.
- Targets exactly ONE shop. Two ways to name that shop:
    * by EXPLICIT ID (WAREHOUSE_IMPORT_TARGET_MODE=shop_id) — the only mode
      allowed against a non-local (production) API, because many accounts have
      a main shop literally named "Asosiy do'kon", so a name is NOT unique.
    * by name (WAREHOUSE_IMPORT_SHOP_NAME) — dev/staging convenience only.
- ID mode verifies, before any write, that:
    * the authenticated token's accountId == WAREHOUSE_IMPORT_ACCOUNT_ID,
    * (super-admin mode) the token role is SUPER_ADMIN,
    * the shop id is one the authenticated account actually OWNS — /api/shops
      is account-scoped, so membership == ownership. The backend independently
      re-checks X-Shop-Id ownership on every call, so a cross-account write is
      impossible even if this guard were bypassed.
- Sets X-Shop-Id to the EXACT target id for every product/category operation, so
  every created/updated row's shop_id is stamped to the target and nowhere else.
- Snapshots per-shop product counts for the whole account BEFORE and AFTER; if any
  NON-target shop's count changed, it STOPS.
- Upserts by product code (Excel "КОД" -> Product.barcode): existing code in the
  shop is UPDATED in place; unknown code is CREATED. Never duplicates.
- Never deletes products that are missing from the Excel.
- Talks only to the API/license URLs you give it (default = localhost dev),
  so it can run against local/staging without ever reaching production.

REQUIRED ENV
------------
  ALLOW_WAREHOUSE_IMPORT=true
  WAREHOUSE_IMPORT_FILE=imports/Товары.xlsx
  WAREHOUSE_IMPORT_USER=<owner/super-admin login>
  WAREHOUSE_IMPORT_PASSWORD=<password>
  # Target — production REQUIRES the id triple; name mode is dev/staging only:
  WAREHOUSE_IMPORT_TARGET_MODE=shop_id          # 'shop_id' (prod) | 'shop_name' (dev)
  WAREHOUSE_IMPORT_SHOP_ID=<exact target shop id>
  WAREHOUSE_IMPORT_ACCOUNT_ID=<exact account id that owns the shop>
  # ...or, dev/staging only:
  WAREHOUSE_IMPORT_SHOP_NAME=Asosiy do'kon
OPTIONAL ENV
  WAREHOUSE_IMPORT_SUPERADMIN=true     # importing as the platform super-admin
  WAREHOUSE_IMPORT_API_URL=http://127.0.0.1:8086
  WAREHOUSE_IMPORT_LICENSE_URL=http://127.0.0.1:19090
  WAREHOUSE_IMPORT_DEFAULT_UNIT=dona
  WAREHOUSE_IMPORT_LIMIT=<N>           # import only first N valid rows (testing)
  WAREHOUSE_IMPORT_PREVIEW=true        # do everything except writes (dry preview)

Mapping decisions (source has no clean field for these — flagged in the report):
- "Ед.из" is a numeric MXIK package code, not a human unit  -> unit defaults to "dona".
- Fractional stock is rounded to the nearest integer (Product.quantity is int).
- Category "Корневая группа" (root group) -> no category (uncategorized).
"""
import base64
import json
import os
import sys
import time
import warnings
from urllib.parse import urlparse

warnings.filterwarnings("ignore")
import openpyxl
import requests

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "_import")
os.makedirs(OUT, exist_ok=True)

CFG = {
    "file": os.environ.get("WAREHOUSE_IMPORT_FILE", os.path.join(HERE, "Товары.xlsx")),
    "shop": os.environ.get("WAREHOUSE_IMPORT_SHOP_NAME", "Asosiy do'kon"),
    # Target locking (production): an explicit shop id + the account that owns it.
    "target_mode": os.environ.get("WAREHOUSE_IMPORT_TARGET_MODE", "shop_name").strip().lower(),
    "shop_id": os.environ.get("WAREHOUSE_IMPORT_SHOP_ID", "").strip(),
    "account_id": os.environ.get("WAREHOUSE_IMPORT_ACCOUNT_ID", "").strip(),
    "superadmin": os.environ.get("WAREHOUSE_IMPORT_SUPERADMIN", "false").lower() == "true",
    "user": os.environ.get("WAREHOUSE_IMPORT_USER", ""),
    "pw": os.environ.get("WAREHOUSE_IMPORT_PASSWORD", ""),
    "api": os.environ.get("WAREHOUSE_IMPORT_API_URL", "http://127.0.0.1:8086").rstrip("/"),
    "lic": os.environ.get("WAREHOUSE_IMPORT_LICENSE_URL", "http://127.0.0.1:19090").rstrip("/"),
    "unit": os.environ.get("WAREHOUSE_IMPORT_DEFAULT_UNIT", "dona"),
    "limit": int(os.environ.get("WAREHOUSE_IMPORT_LIMIT", "0") or "0"),
    "preview": os.environ.get("WAREHOUSE_IMPORT_PREVIEW", "false").lower() == "true",
    # Stay under the backend's per-caller rate limit (default 600/min). 0 = no
    # throttle (use only when the server limit is raised/disabled for staging).
    "rate_per_min": int(os.environ.get("WAREHOUSE_IMPORT_RATE_PER_MIN", "540") or "0"),
    "retries": int(os.environ.get("WAREHOUSE_IMPORT_RETRIES", "6")),
}
ROOT_CATEGORY = "Корневая группа"  # treated as "no category"

# Hosts considered "local" — a non-local API host means production, where the
# stricter id-locking rules below are mandatory.
LOCAL_HOSTS = {"127.0.0.1", "localhost", "0.0.0.0", "::1", ""}


def is_production(api_url):
    """True when the API host is not local — production safety rules then apply."""
    host = (urlparse(api_url).hostname or "").lower()
    return host not in LOCAL_HOSTS


def decode_jwt_claims(token):
    """Read (WITHOUT verifying) the JWT payload so we can assert the token's
    accountId / role match the explicit target before sending any write. The
    signature is still verified server-side on every call — this is only a
    client-side pre-flight guard, never a trust boundary."""
    try:
        payload = token.split(".")[1]
        payload += "=" * (-len(payload) % 4)  # restore base64url padding
        return json.loads(base64.urlsafe_b64decode(payload.encode("ascii")))
    except Exception:
        return {}


def verify_target(claims, shops, shop_id, account_id, superadmin):
    """Resolve the target shop strictly by id, enforcing every guard:
      - the token's accountId must equal the explicit account_id (when given),
      - super-admin mode requires a SUPER_ADMIN token,
      - the shop id must be one the authenticated account actually owns
        (/api/shops is account-scoped, so membership == ownership).
    Returns the matched shop dict, or raises SystemExit with a clear reason."""
    tok_acct = claims.get("accountId")
    tok_role = claims.get("role")
    if account_id is not None and tok_acct != account_id:
        raise SystemExit(
            f"ABORT: authenticated accountId={tok_acct} does not match "
            f"WAREHOUSE_IMPORT_ACCOUNT_ID={account_id} — wrong credentials for this target.")
    if superadmin and tok_role != "SUPER_ADMIN":
        raise SystemExit(
            f"ABORT: WAREHOUSE_IMPORT_SUPERADMIN=true but the token role is {tok_role!r}, "
            f"not SUPER_ADMIN.")
    match = [s for s in shops if int(s.get("id")) == int(shop_id)]
    if not match:
        avail = [s.get("id") for s in shops]
        raise SystemExit(
            f"ABORT: shop_id={shop_id} is not owned by the authenticated account "
            f"(accountId={tok_acct}). Shops it owns: {avail}. "
            f"If you are super-admin and the target shop belongs to ANOTHER account, "
            f"super-admin CANNOT create products there (the backend rejects a cross-account "
            f"X-Shop-Id) — log in as that account's OWNER instead.")
    return match[0]


def assert_only_target_changed(before, after, target_id):
    """Guard: every non-target shop's product count must be identical before and
    after. Returns a list of offending [shop_id, before, after]; empty == safe."""
    drift = []
    for sid, b in before.items():
        if int(sid) == int(target_id):
            continue
        a = after.get(sid)
        if a != b:
            drift.append([sid, b, a])
    return drift

HEADER_ALIASES = {
    "code": ["код", "kod", "code"], "name": ["товар", "наименование", "name"],
    "category": ["категория", "kategoriya"], "qty": ["остаток", "qoldiq", "stock"],
    "cost": ["себестоимость", "tannarx", "cost"], "sell": ["цена", "narx", "sell", "price"],
    "mxik": ["икпу", "ikpu", "mxik"], "unit": ["ед.из", "ед.изм", "unit"],
    "minstock": ["мин.остаток", "min", "minimal"],
}


def norm(s):
    return ("" if s is None else str(s)).strip().lower()


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s) if s else None
    except ValueError:
        return None


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    flat = {a: f for f, al in HEADER_ALIASES.items() for a in al}
    hidx, colmap, hits = 0, {}, -1
    for i, row in enumerate(rows[:25]):
        cm = {}
        for j, c in enumerate(row):
            k = norm(c)
            if k in flat and flat[k] not in cm:
                cm[flat[k]] = j
        if len(cm) > hits:
            hits, hidx, colmap = len(cm), i, cm
    valid, invalid, frac, negative_corrected = [], [], 0, []

    def cell(row, f):
        j = colmap.get(f)
        return row[j] if j is not None and j < len(row) else None

    for n, row in enumerate(rows[hidx + 1:]):
        excel_row = hidx + 1 + n + 1
        code = "" if cell(row, "code") is None else str(cell(row, "code")).strip()
        name = "" if cell(row, "name") is None else str(cell(row, "name")).strip()
        if code == "" and name == "":
            continue
        qty, cost, sell = to_num(cell(row, "qty")), to_num(cell(row, "cost")), to_num(cell(row, "sell"))
        minst = to_num(cell(row, "minstock"))
        mxik = cell(row, "mxik"); mxik = "" if mxik is None else str(mxik).strip()
        cat = cell(row, "category"); cat = "" if cat is None else str(cat).strip()
        problems = []
        if code == "": problems.append("code missing")
        if name == "": problems.append("name missing")
        if qty is None: problems.append("qty not numeric")
        if cost is None: problems.append("cost not numeric")
        elif cost < 0: problems.append("cost negative")
        if sell is None: problems.append("sell not numeric")
        elif sell < 0: problems.append("sell negative")
        rec = {"excel_row": excel_row, "code": code, "name": name, "category": cat,
               "qty": qty, "qty_original": qty, "cost": cost, "sell": sell,
               "mxik": mxik, "minstock": minst, "negative_corrected": False,
               "problems": problems}
        if problems:
            invalid.append(rec)
            continue
        # Negative stock is a data artefact (oversold rows). Correct it to 0 and
        # REPORT it (never silently hidden) — the row is still imported.
        if qty < 0:
            rec["negative_corrected"] = True
            rec["qty"] = 0.0
            negative_corrected.append(rec)
        if rec["qty"] != round(rec["qty"]):
            frac += 1
        valid.append(rec)
    wb.close()
    return valid, invalid, frac, hidx + 1, colmap, negative_corrected


class Api:
    def __init__(self, cfg):
        self.cfg = cfg
        self.s = requests.Session()
        self.token = None
        self.shop_id = None
        self._interval = 60.0 / cfg["rate_per_min"] if cfg["rate_per_min"] > 0 else 0.0
        self._last = 0.0

    def _throttle(self):
        if self._interval:
            dt = time.time() - self._last
            if dt < self._interval:
                time.sleep(self._interval - dt)
            self._last = time.time()

    def _send(self, method, url, **kw):
        """Request with client-side throttle + 429 retry/backoff."""
        kw.setdefault("timeout", 30)
        resp = None
        for attempt in range(self.cfg["retries"]):
            self._throttle()
            resp = self.s.request(method, url, **kw)
            if resp.status_code == 429:
                time.sleep(min(30.0, 2.0 ** attempt + 0.5))
                continue
            return resp
        return resp

    def login(self):
        r = self.s.post(f"{self.cfg['lic']}/api/auth/login",
                        json={"username": self.cfg["user"], "password": self.cfg["pw"]},
                        timeout=15)
        r.raise_for_status()
        self.token = r.json()["token"]
        self.s.headers.update({"Authorization": f"Bearer {self.token}"})

    def _h(self):
        return {"X-Shop-Id": str(self.shop_id)} if self.shop_id else {}

    def jwt_claims(self):
        return decode_jwt_claims(self.token or "")

    def list_shops(self):
        """All shops the authenticated account owns (the API is account-scoped)."""
        r = self.s.get(f"{self.cfg['api']}/api/shops", timeout=15)
        r.raise_for_status()
        return r.json()

    def count_products_for(self, shop_id):
        """Number of products in one shop of the authenticated account. Reads are
        allowed for any owned shop; used to snapshot the count guard."""
        r = self.s.get(f"{self.cfg['api']}/api/products",
                       headers={"X-Shop-Id": str(shop_id)}, timeout=120)
        r.raise_for_status()
        return len(r.json())

    def resolve_shop_by_id(self, shop_id, account_id, superadmin):
        """Production path: lock the target to an exact id, after verifying the
        token's account/role and that the account owns the shop. Returns
        (matched_shop, [all owned shop ids])."""
        shops = self.list_shops()
        shop = verify_target(self.jwt_claims(), shops, shop_id, account_id, superadmin)
        self.shop_id = int(shop_id)
        return shop, [int(s["id"]) for s in shops]

    def resolve_shop(self, name):
        r = self.s.get(f"{self.cfg['api']}/api/shops", timeout=15)
        r.raise_for_status()
        shops = r.json()
        match = [s for s in shops if str(s.get("name", "")).strip() == name.strip()]
        if not match:
            avail = [s.get("name") for s in shops]
            raise SystemExit(f"ABORT: target shop {name!r} not found. Available: {avail}")
        if len(match) > 1:
            raise SystemExit(f"ABORT: {len(match)} shops named {name!r} — ambiguous.")
        self.shop_id = match[0]["id"]
        return match[0]

    def existing_products(self):
        r = self.s.get(f"{self.cfg['api']}/api/products", headers=self._h(), timeout=60)
        r.raise_for_status()
        out = {}
        for p in r.json():
            bc = (p.get("barcode") or "").strip()
            if bc:
                out[bc] = p
        return out

    def create(self, payload):
        r = self._send("POST", f"{self.cfg['api']}/api/products", headers=self._h(), json=payload)
        r.raise_for_status()
        return r.json()

    def update(self, pid, payload):
        r = self._send("PUT", f"{self.cfg['api']}/api/products/{pid}", headers=self._h(), json=payload)
        r.raise_for_status()
        return r.json()

    def stocktake(self, counts):
        r = self._send("POST", f"{self.cfg['api']}/api/products/stocktake", headers=self._h(),
                       json={"counts": counts}, timeout=120)
        r.raise_for_status()
        return r.json()

    def categories(self):
        r = self.s.get(f"{self.cfg['api']}/api/categories", headers=self._h(), timeout=30)
        r.raise_for_status()
        return r.json()

    def create_category(self, name):
        r = self._send("POST", f"{self.cfg['api']}/api/categories", headers=self._h(),
                       json={"name": name})
        r.raise_for_status()
        return r.json()

    def resolve_categories(self, names):
        """Map each category name -> id, creating any that don't exist yet."""
        existing = {c["name"].strip().lower(): c["id"] for c in self.categories()}
        out = {}
        for name in names:
            key = name.strip().lower()
            if key in existing:
                out[name] = existing[key]
            else:
                cid = self.create_category(name)["id"]
                existing[key] = cid
                out[name] = cid
        return out


def payload_for(rec, cfg, cat_map=None):
    cat = rec["category"]
    is_real = cat not in ("", ROOT_CATEGORY)
    # Send an explicit categoryId so BOTH create and update set it (the backend's
    # categoryName resolve-or-create only fires on create, so an idempotent
    # re-import would otherwise null the category on update).
    category_id = (cat_map or {}).get(cat) if is_real else None
    return {
        "name": rec["name"],
        "barcode": rec["code"],
        "purchasePrice": round(rec["cost"], 2),
        "salePrice": round(rec["sell"], 2),
        "quantity": int(round(rec["qty"])),
        "categoryId": category_id,
        "categoryName": cat if (is_real and category_id is None) else None,
        "lowStockThreshold": int(round(rec["minstock"])) if rec["minstock"] is not None else 0,
        "mxikCode": rec["mxik"] or None,
        "unit": cfg["unit"],
        "requiresImei": False,
    }


def main():
    if os.environ.get("ALLOW_WAREHOUSE_IMPORT", "").lower() != "true":
        raise SystemExit("REFUSING: set ALLOW_WAREHOUSE_IMPORT=true to run the importer.")
    if not CFG["user"] or not CFG["pw"]:
        raise SystemExit("REFUSING: WAREHOUSE_IMPORT_USER / WAREHOUSE_IMPORT_PASSWORD required.")

    prod = is_production(CFG["api"])
    mode = CFG["target_mode"]
    shop_id = int(CFG["shop_id"]) if CFG["shop_id"] else None
    account_id = int(CFG["account_id"]) if CFG["account_id"] else None

    # --- Target-locking guards (production must address the shop by exact id) ---
    if prod and mode != "shop_id":
        raise SystemExit(
            "REFUSING: production import must target an explicit shop id. Set "
            "WAREHOUSE_IMPORT_TARGET_MODE=shop_id, WAREHOUSE_IMPORT_SHOP_ID and "
            "WAREHOUSE_IMPORT_ACCOUNT_ID. Name-only resolution is forbidden against a "
            "non-local API — many accounts share the shop name \"Asosiy do'kon\".")
    if prod and (shop_id is None or account_id is None):
        raise SystemExit(
            "REFUSING: production import requires BOTH WAREHOUSE_IMPORT_SHOP_ID and "
            "WAREHOUSE_IMPORT_ACCOUNT_ID (exact integers).")
    if CFG["superadmin"] and (shop_id is None or account_id is None):
        raise SystemExit(
            "REFUSING: WAREHOUSE_IMPORT_SUPERADMIN=true requires both "
            "WAREHOUSE_IMPORT_SHOP_ID and WAREHOUSE_IMPORT_ACCOUNT_ID.")
    if mode == "shop_id" and shop_id is None:
        raise SystemExit(
            "REFUSING: WAREHOUSE_IMPORT_TARGET_MODE=shop_id requires WAREHOUSE_IMPORT_SHOP_ID.")

    print(f"File    : {CFG['file']}")
    print(f"Target  : mode={mode} shop_id={shop_id} account_id={account_id} "
          f"superadmin={CFG['superadmin']}  (name={CFG['shop']!r})")
    print(f"API     : {CFG['api']}   License: {CFG['lic']}   production={prod}")
    print(f"Preview : {CFG['preview']}   Limit: {CFG['limit'] or 'all'}")
    valid, invalid, frac, header_row, colmap, negative_corrected = parse_excel(CFG["file"])
    print(f"Parsed  : valid={len(valid)} invalid={len(invalid)} "
          f"negative_stock_corrected_to_zero={len(negative_corrected)} "
          f"fractional_qty_rounded={frac} header_row={header_row}")
    if CFG["limit"]:
        valid = valid[:CFG["limit"]]

    api = Api(CFG)
    api.login()
    if mode == "shop_id":
        shop, account_shop_ids = api.resolve_shop_by_id(shop_id, account_id, CFG["superadmin"])
    else:
        # Name mode is dev/staging only (the production guard above already
        # refused this path against a non-local API).
        shop = api.resolve_shop(CFG["shop"])
        account_shop_ids = [int(s["id"]) for s in api.list_shops()]
    target_id = int(shop["id"])
    print(f"Shop OK : id={target_id} name={shop.get('name')!r} main={shop.get('main')}")

    # Snapshot every shop in this account BEFORE any write. The target may grow;
    # every other shop must be byte-for-byte unchanged afterwards.
    baseline = {sid: api.count_products_for(sid) for sid in account_shop_ids}
    print(f"Baseline product counts (this account's shops): {baseline}")

    existing = api.existing_products()
    print(f"Existing products in target shop: {len(existing)}")

    # Pre-resolve real categories to ids (create missing) so create AND update
    # both set categoryId. "Корневая группа" (root) stays uncategorised.
    cat_names = sorted({r["category"] for r in valid
                        if r["category"] not in ("", ROOT_CATEGORY)})
    # Preview is strictly read-only: don't create categories on a dry-run.
    cat_map = ({} if CFG["preview"]
               else api.resolve_categories(cat_names) if cat_names else {})
    print(f"Categories resolved/created: {len(cat_map)} -> {list(cat_map.keys())}")

    created = updated = skipped = 0
    errors = []
    qty_fixes = []
    for i, rec in enumerate(valid):
        payload = payload_for(rec, CFG, cat_map)
        try:
            if rec["code"] in existing:
                pid = existing[rec["code"]]["id"]
                if not CFG["preview"]:
                    api.update(pid, payload)
                    cur = int(existing[rec["code"]].get("quantity") or 0)
                    if cur != payload["quantity"]:
                        qty_fixes.append({"productId": pid, "actual": payload["quantity"]})
                updated += 1
            else:
                if not CFG["preview"]:
                    api.create(payload)
                created += 1
        except requests.HTTPError as e:
            body = e.response.text[:200] if e.response is not None else str(e)
            errors.append({"code": rec["code"], "name": rec["name"], "error": body})
            skipped += 1
        if (i + 1) % 500 == 0:
            print(f"  ...{i + 1}/{len(valid)}  created={created} updated={updated} errors={len(errors)}")

    # Correct stock on updated products whose qty drifted (re-import path).
    if qty_fixes and not CFG["preview"]:
        for k in range(0, len(qty_fixes), 500):
            api.stocktake(qty_fixes[k:k + 500])

    # --- Count guard: re-snapshot and prove no NON-target shop changed. -------
    # Preview writes nothing, so after == baseline by construction; we still
    # record it. On a real import this is the stop-gate: any drift => STOP.
    after = dict(baseline)
    if not CFG["preview"]:
        after = {sid: api.count_products_for(sid) for sid in account_shop_ids}
    drift = assert_only_target_changed(baseline, after, target_id)
    if drift:
        print("!!! NON-TARGET SHOP DRIFT DETECTED (shop_id, before, after):", drift)
        raise SystemExit(
            f"STOP: {len(drift)} non-target shop(s) changed — investigate before any further "
            f"action. Drift: {drift}")

    result = {
        "shop": {"id": shop["id"], "name": shop.get("name"), "main": shop.get("main")},
        "target_mode": mode, "superadmin": CFG["superadmin"],
        "account_id": account_id, "target_shop_id": target_id,
        "shop_count_guard": {
            "account_shop_ids": account_shop_ids,
            "before": baseline, "after": after,
            "non_target_drift": drift,
        },
        "valid_rows": len(valid), "invalid_rows": len(invalid),
        "created": created, "updated": updated, "skipped_errors": skipped,
        "stock_corrections": len(qty_fixes), "fractional_qty_rounded": frac,
        "negative_stock_corrected_to_zero": len(negative_corrected),
        "negative_stock_rows": [
            {"excel_row": r["excel_row"], "code": r["code"], "name": r["name"],
             "original_qty": r["qty_original"], "imported_qty": 0}
            for r in negative_corrected
        ],
        "invalid_row_details": [
            {"excel_row": r["excel_row"], "code": r["code"], "name": r["name"],
             "problems": r["problems"]} for r in invalid
        ],
        "preview": CFG["preview"], "errors": errors[:50],
    }
    with open(os.path.join(OUT, "result.json"), "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print("-" * 60)
    _verbose = {"errors", "negative_stock_rows", "invalid_row_details"}
    print(json.dumps({k: v for k, v in result.items() if k not in _verbose},
                     ensure_ascii=False, indent=2))
    if errors:
        print(f"ERRORS ({len(errors)}), first 5:")
        for e in errors[:5]:
            print("   ", e)
    print(f"\nResult written to {os.path.join(OUT, 'result.json')}")


if __name__ == "__main__":
    main()
