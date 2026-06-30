"""
Warehouse Excel import — DRY RUN ONLY (no DB, no API, read-only).

Detects the header row, maps columns by header name, counts/validates rows,
finds duplicate codes, computes stock/cost/sell totals, and lists invalid rows.
Writes evidence files next to the workbook. Does NOT import anything.

Usage:  PYTHONUTF8=1 python imports/dryrun_warehouse_import.py
"""
import json
import os
from collections import Counter, defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Товары.xlsx")
OUT = os.path.join(HERE, "_dryrun")
os.makedirs(OUT, exist_ok=True)

# Canonical field -> accepted header aliases (lower-cased, stripped).
HEADER_ALIASES = {
    "code":     ["код", "kod", "code", "артикул"],
    "name":     ["товар", "наименование", "название", "mahsulot", "name"],
    "category": ["категория", "kategoriya", "category", "группа"],
    "qty":      ["остаток", "qoldiq", "количество", "stock", "soni"],
    "cost":     ["себестоимость", "tannarx", "закуп", "cost", "purchase"],
    "markup":   ["наценко", "наценка", "markup", "ustama"],
    "sell":     ["цена", "narx", "sell", "price", "sotuv"],
    "mxik":     ["икпу", "ikpu", "mxik", "мхик"],
    "unit":     ["ед.из", "ед.изм", "edizm", "unit", "birlik", "ед"],
    "minstock": ["мин.остаток", "min", "минимальный", "minimal", "min.qoldiq"],
}
REQUIRED_COLS = ["code", "name", "category", "qty", "cost", "sell", "mxik", "unit", "minstock"]


def norm(s):
    return ("" if s is None else str(s)).strip().lower()


def to_num(v):
    """Return float or None if not a usable number."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # 1) Detect header row: the row whose cells match the most known aliases.
    flat_aliases = {a: f for f, al in HEADER_ALIASES.items() for a in al}
    best_idx, best_map, best_hits = None, None, -1
    for i, row in enumerate(rows[:25]):
        colmap = {}
        for j, cell in enumerate(row):
            key = norm(cell)
            if key in flat_aliases and flat_aliases[key] not in colmap:
                colmap[flat_aliases[key]] = j
        if len(colmap) > best_hits:
            best_hits, best_idx, best_map = len(colmap), i, colmap
    header_idx, colmap = best_idx, best_map

    missing_required = [c for c in REQUIRED_COLS if c not in colmap]

    # 2) Iterate data rows (after the header). A data row has a non-empty code.
    data = rows[header_idx + 1:]
    valid, invalid, seen_codes = [], [], Counter()
    cat_counter, unit_counter = Counter(), Counter()
    total_qty = total_cost_value = total_sell_value = 0.0
    n_missing_mxik = n_missing_unit = n_zero_qty = 0
    blank_rows = 0

    def cell(row, field):
        idx = colmap.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    for n, row in enumerate(data):
        excel_row = header_idx + 1 + n + 1  # 1-based Excel row number
        code_raw = cell(row, "code")
        name_raw = cell(row, "name")
        # Skip fully blank rows / trailing footer rows.
        if (code_raw is None or str(code_raw).strip() == "") and \
           (name_raw is None or str(name_raw).strip() == ""):
            blank_rows += 1
            continue

        code = "" if code_raw is None else str(code_raw).strip()
        name = "" if name_raw is None else str(name_raw).strip()
        qty = to_num(cell(row, "qty"))
        cost = to_num(cell(row, "cost"))
        sell = to_num(cell(row, "sell"))
        minst = to_num(cell(row, "minstock"))
        mxik = cell(row, "mxik")
        mxik = "" if mxik is None else str(mxik).strip()
        unit = cell(row, "unit")
        unit = "" if unit is None else str(unit).strip()
        category = cell(row, "category")
        category = "" if category is None else str(category).strip()

        problems = []
        if code == "":
            problems.append("code missing")
        if name == "":
            problems.append("name missing")
        if qty is None:
            problems.append("qty not numeric")
        elif qty < 0:
            problems.append("qty negative")
        if cost is None:
            problems.append("cost not numeric")
        elif cost < 0:
            problems.append("cost negative")
        if sell is None:
            problems.append("sell not numeric")
        elif sell < 0:
            problems.append("sell negative")

        rec = {"excel_row": excel_row, "code": code, "name": name,
               "category": category, "qty": qty, "cost": cost, "sell": sell,
               "mxik": mxik, "unit": unit, "minstock": minst,
               "problems": problems}

        if problems:
            invalid.append(rec)
            continue

        seen_codes[code] += 1
        cat_counter[category or "(bo'sh)"] += 1
        unit_counter[unit or "(bo'sh)"] += 1
        if mxik == "":
            n_missing_mxik += 1
        if unit == "" or unit == "0":
            n_missing_unit += 1
        if qty == 0:
            n_zero_qty += 1
        total_qty += qty
        total_cost_value += qty * cost
        total_sell_value += qty * sell
        valid.append(rec)

    duplicates = {c: k for c, k in seen_codes.items() if k > 1}
    dup_rows = [v for v in valid if v["code"] in duplicates]

    # ---- Console report ----
    print("=" * 70)
    print("WAREHOUSE EXCEL IMPORT — DRY RUN")
    print("=" * 70)
    print(f"File              : {XLSX}")
    print(f"Sheet             : {ws.title}   (sheets={wb.sheetnames})")
    print(f"Header row        : Excel row {header_idx + 1} (0-based index {header_idx})")
    print(f"Column mapping    :")
    for f in REQUIRED_COLS + ["markup"]:
        j = colmap.get(f)
        col_letter = openpyxl.utils.get_column_letter(j + 1) if j is not None else "—"
        print(f"    {f:<9} -> col {col_letter} (idx {j})")
    print(f"Missing required  : {missing_required or 'NONE ✓'}")
    print("-" * 70)
    print(f"Rows below header : {len(data)}")
    print(f"Blank/footer rows : {blank_rows}")
    print(f"VALID products    : {len(valid)}")
    print(f"INVALID rows      : {len(invalid)}")
    print(f"Distinct codes    : {len(seen_codes)}")
    print(f"Duplicate codes   : {len(duplicates)}  (rows involved: {len(dup_rows)})")
    print("-" * 70)
    print(f"TOTAL stock (qty)        : {total_qty:,.2f}")
    print(f"TOTAL stock COST value   : {total_cost_value:,.2f}")
    print(f"TOTAL stock SELLING value: {total_sell_value:,.2f}")
    print(f"Potential gross margin   : {total_sell_value - total_cost_value:,.2f}")
    print("-" * 70)
    print(f"Products w/ EMPTY mxik    : {n_missing_mxik}")
    print(f"Products w/ EMPTY/0 unit  : {n_missing_unit}")
    print(f"Products w/ ZERO qty      : {n_zero_qty}")
    print("-" * 70)
    print(f"Distinct categories: {len(cat_counter)}")
    for c, k in cat_counter.most_common(15):
        print(f"    {k:>5}  {c}")
    print(f"Distinct unit codes: {len(unit_counter)}")
    for u, k in unit_counter.most_common(12):
        print(f"    {k:>5}  {u}")
    print("-" * 70)
    if duplicates:
        print("DUPLICATE CODES (code -> count):")
        for c, k in sorted(duplicates.items(), key=lambda x: -x[1])[:30]:
            names = [v["name"] for v in dup_rows if v["code"] == c]
            print(f"    {c} x{k}: {names}")
    if invalid:
        print(f"\nFIRST 25 INVALID ROWS:")
        for r in invalid[:25]:
            print(f"    Excel R{r['excel_row']}: code={r['code']!r} name={r['name'][:30]!r} -> {r['problems']}")

    # ---- Evidence files ----
    summary = {
        "file": XLSX, "sheet": ws.title, "header_excel_row": header_idx + 1,
        "column_mapping": {f: colmap.get(f) for f in REQUIRED_COLS + ["markup"]},
        "missing_required": missing_required,
        "rows_below_header": len(data), "blank_rows": blank_rows,
        "valid": len(valid), "invalid": len(invalid),
        "distinct_codes": len(seen_codes), "duplicate_codes": len(duplicates),
        "duplicate_rows": len(dup_rows),
        "total_qty": total_qty, "total_cost_value": total_cost_value,
        "total_sell_value": total_sell_value,
        "empty_mxik": n_missing_mxik, "empty_or_zero_unit": n_missing_unit,
        "zero_qty": n_zero_qty,
        "distinct_categories": len(cat_counter),
        "categories": dict(cat_counter),
        "distinct_units": len(unit_counter),
    }
    with open(os.path.join(OUT, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, "invalid_rows.json"), "w", encoding="utf-8") as f:
        json.dump(invalid, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, "duplicates.json"), "w", encoding="utf-8") as f:
        json.dump({c: [v for v in dup_rows if v["code"] == c] for c in duplicates},
                  f, ensure_ascii=False, indent=2)
    print(f"\nEvidence written to: {OUT}")
    wb.close()


if __name__ == "__main__":
    main()
